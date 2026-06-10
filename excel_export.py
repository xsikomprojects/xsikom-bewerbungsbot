import sqlite3
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from database import DB_NAME


def excel_export():
    """Exportiert alle Daten in eine Excel-Datei."""

    conn = sqlite3.connect(DB_NAME)
    c    = conn.cursor()

    # Daten laden
    c.execute("""
        SELECT id, titel, firma, standort, email,
               url, quelle, datum, status
        FROM stellen ORDER BY id DESC
    """)
    stellen = c.fetchall()

    c.execute("""
        SELECT id, firma, position, email,
               status, datum
        FROM bewerbungen ORDER BY id DESC
    """)
    bewerbungen = c.fetchall()

    conn.close()

    # Excel erstellen
    wb = Workbook()

    # ── FARBEN ────────────────────────────────────────
    blau     = PatternFill("solid", fgColor="00467F")
    hellblau = PatternFill("solid", fgColor="DCE6F1")
    gruen    = PatternFill("solid", fgColor="E2EFDA")
    gelb     = PatternFill("solid", fgColor="FFF2CC")
    rot      = PatternFill("solid", fgColor="FFE0E0")

    weiss  = Font(color="FFFFFF", bold=True, size=11)
    dunkel = Font(color="00467F", bold=True, size=10)
    normal = Font(size=9)

    rand = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    zentriert = Alignment(horizontal="center")

    # ============================================================
    # BLATT 1: STELLEN
    # ============================================================
    ws1       = wb.active
    ws1.title = "Gefundene Stellen"

    # Titel Zeile 1
    ws1["A1"] = "IT-PRAKTIKUM BEWERBUNGSBOT - Komi Tevi"
    ws1["A1"].fill      = blau
    ws1["A1"].font      = weiss
    ws1["A1"].alignment = zentriert
    ws1.merge_cells("A1:I1")

    # Datum Zeile 2
    ws1["A2"] = f"Exportiert am: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws1["A2"].fill      = hellblau
    ws1["A2"].font      = dunkel
    ws1["A2"].alignment = zentriert
    ws1.merge_cells("A2:I2")

    # Header Zeile 3
    headers1 = [
        "ID", "Titel", "Firma", "Standort",
        "E-Mail", "URL", "Quelle", "Datum", "Status"
    ]
    for col, h in enumerate(headers1, 1):
        zelle            = ws1.cell(row=3, column=col, value=h)
        zelle.fill       = blau
        zelle.font       = weiss
        zelle.alignment  = zentriert
        zelle.border     = rand

    # Daten ab Zeile 4
    for row, s in enumerate(stellen, 4):
        for col, wert in enumerate(s, 1):
            zelle        = ws1.cell(row=row, column=col, value=wert)
            zelle.border = rand
            zelle.font   = normal

            # Farbe je nach Status
            if len(s) > 8:
                status = str(s[8])
                if status == "neu":
                    zelle.fill = hellblau
                elif status == "beworben":
                    zelle.fill = gruen

    # Spaltenbreiten BLATT 1
    breiten1 = {
        "A": 5,
        "B": 30,
        "C": 25,
        "D": 15,
        "E": 30,
        "F": 35,
        "G": 12,
        "H": 16,
        "I": 10,
    }
    for spalte, breite in breiten1.items():
        ws1.column_dimensions[spalte].width = breite

    # ============================================================
    # BLATT 2: BEWERBUNGEN
    # ============================================================
    ws2       = wb.create_sheet("Bewerbungen")

    ws2["A1"] = "BEWERBUNGEN - Komi Tevi"
    ws2["A1"].fill      = blau
    ws2["A1"].font      = weiss
    ws2["A1"].alignment = zentriert
    ws2.merge_cells("A1:F1")

    ws2["A2"] = f"Exportiert am: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws2["A2"].fill      = hellblau
    ws2["A2"].font      = dunkel
    ws2["A2"].alignment = zentriert
    ws2.merge_cells("A2:F2")

    headers2 = ["ID", "Firma", "Position", "E-Mail", "Status", "Datum"]
    for col, h in enumerate(headers2, 1):
        zelle           = ws2.cell(row=3, column=col, value=h)
        zelle.fill      = blau
        zelle.font      = weiss
        zelle.alignment = zentriert
        zelle.border    = rand

    for row, b in enumerate(bewerbungen, 4):
        for col, wert in enumerate(b, 1):
            zelle        = ws2.cell(row=row, column=col, value=wert)
            zelle.border = rand
            zelle.font   = normal

            if len(b) > 4:
                status = str(b[4])
                if status == "gesendet":
                    zelle.fill = gruen
                elif status == "trockenlauf":
                    zelle.fill = gelb
                elif status == "fehler":
                    zelle.fill = rot

    # Spaltenbreiten BLATT 2
    breiten2 = {
        "A": 5,
        "B": 25,
        "C": 25,
        "D": 30,
        "E": 12,
        "F": 16,
    }
    for spalte, breite in breiten2.items():
        ws2.column_dimensions[spalte].width = breite

    # ============================================================
    # BLATT 3: STATISTIKEN
    # ============================================================
    ws3       = wb.create_sheet("Statistiken")

    ws3["A1"] = "STATISTIKEN - Komi Tevi"
    ws3["A1"].fill      = blau
    ws3["A1"].font      = weiss
    ws3["A1"].alignment = zentriert
    ws3.merge_cells("A1:C1")

    ws3["A2"] = f"Stand: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws3["A2"].fill      = hellblau
    ws3["A2"].font      = dunkel
    ws3.merge_cells("A2:C2")

    # Statistik Daten
    bew_gesendet  = sum(1 for b in bewerbungen if b[4] == "gesendet")
    bew_trockenlauf = sum(1 for b in bewerbungen if b[4] == "trockenlauf")
    bew_fehler    = sum(1 for b in bewerbungen if b[4] == "fehler")

    stats_daten = [
        ("Stellen gesamt",       len(stellen)),
        ("Bewerbungen gesamt",   len(bewerbungen)),
        ("Bewerbungen gesendet", bew_gesendet),
        ("Testlaeufe",           bew_trockenlauf),
        ("Fehler",               bew_fehler),
    ]

    for row, (label, wert) in enumerate(stats_daten, 3):
        l_zelle        = ws3.cell(row=row, column=1, value=label)
        l_zelle.font   = dunkel
        l_zelle.border = rand

        w_zelle        = ws3.cell(row=row, column=2, value=wert)
        w_zelle.font   = normal
        w_zelle.border = rand

        if label == "Bewerbungen gesendet":
            w_zelle.fill = gruen
        elif label == "Fehler":
            w_zelle.fill = rot

    ws3.column_dimensions["A"].width = 25
    ws3.column_dimensions["B"].width = 15
    ws3.column_dimensions["C"].width = 15

    # ============================================================
    # SPEICHERN
    # ============================================================
    datum     = datetime.now().strftime("%Y-%m-%d_%H-%M")
    dateiname = f"Bewerbungen_Komi_Tevi_{datum}.xlsx"
    pfad      = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        dateiname
    )

    wb.save(pfad)

    print(f"\n  {'='*50}")
    print(f"  EXCEL EXPORT ERFOLGREICH!")
    print(f"  {'='*50}")
    print(f"  Datei       : {dateiname}")
    print(f"  Stellen     : {len(stellen)}")
    print(f"  Bewerbungen : {len(bewerbungen)}")
    print(f"  {'='*50}")
    print(f"\n  Pfad: {pfad}")

    # Excel automatisch öffnen
    try:
        os.startfile(pfad)
        print("  Excel wird geöffnet...")
    except Exception:
        pass

    return pfad


if __name__ == "__main__":
    excel_export()